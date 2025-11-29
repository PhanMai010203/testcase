"""
Android Test Case Generator
A PyQt6 application that generates test cases from images or documents using Gemini API.
"""

import sys
import os
import json
from datetime import datetime
from pathlib import Path
from typing import Optional, List

from PyQt6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QLabel, QPushButton, QTextEdit, QFileDialog, QFrame, QScrollArea,
    QSplitter, QTabWidget, QProgressBar, QMessageBox, QLineEdit,
    QComboBox, QCheckBox, QGroupBox, QGridLayout, QSizePolicy,
    QButtonGroup, QRadioButton, QStackedWidget, QSpacerItem
)
from PyQt6.QtCore import Qt, QThread, pyqtSignal, QSize, QPropertyAnimation, QEasingCurve, QTimer
from PyQt6.QtGui import (
    QPixmap, QFont, QColor, QPalette, QDragEnterEvent, QDropEvent,
    QIcon, QPainter, QLinearGradient, QBrush, QPen
)

import google.generativeai as genai
from PIL import Image
import PyPDF2
import docx
import re
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


# ═══════════════════════════════════════════════════════════════════════════════
# STYLING - Cyberpunk/Neon Theme
# ═══════════════════════════════════════════════════════════════════════════════

STYLESHEET = """
* {
    font-family: 'JetBrains Mono', 'Fira Code', 'SF Mono', monospace;
}

QMainWindow {
    background: qlineargradient(x1:0, y1:0, x2:1, y2:1,
        stop:0 #0a0a0f, stop:0.5 #12121a, stop:1 #0d0d14);
}

QWidget {
    color: #e0e0e0;
    background: transparent;
}

QLabel {
    color: #e0e0e0;
    font-size: 13px;
}

QLabel#title {
    font-size: 28px;
    font-weight: 700;
    color: #00ffd5;
    letter-spacing: 2px;
}

QLabel#subtitle {
    font-size: 12px;
    color: #7a7a8c;
    letter-spacing: 1px;
}

QLabel#sectionTitle {
    font-size: 16px;
    font-weight: 600;
    color: #ff6b9d;
    border-bottom: 2px solid #ff6b9d;
    padding-bottom: 8px;
    margin-bottom: 12px;
}

QLabel#modeTitle {
    font-size: 14px;
    font-weight: 600;
    color: #00ffd5;
    padding: 4px 0;
}

QLabel#statusActive {
    color: #00ffd5;
    font-size: 12px;
    font-weight: 600;
    padding: 6px 12px;
    background: rgba(0, 255, 213, 0.15);
    border-radius: 4px;
    border: 1px solid rgba(0, 255, 213, 0.3);
}

QLabel#statusInactive {
    color: #7a7a8c;
    font-size: 12px;
    padding: 6px 12px;
    background: rgba(45, 45, 68, 0.5);
    border-radius: 4px;
}

QPushButton {
    background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
        stop:0 #1a1a2e, stop:1 #16213e);
    border: 1px solid #00ffd5;
    border-radius: 8px;
    padding: 12px 24px;
    font-size: 13px;
    font-weight: 600;
    color: #00ffd5;
    min-height: 20px;
}

QPushButton:hover {
    background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
        stop:0 #00ffd5, stop:1 #00d4aa);
    color: #0a0a0f;
    border: 1px solid #00ffd5;
}

QPushButton:pressed {
    background: #00aa88;
}

QPushButton:disabled {
    background: #1a1a2e;
    border: 1px solid #2d2d44;
    color: #4a4a5c;
}

QPushButton#generateBtn {
    background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
        stop:0 #ff6b9d, stop:1 #c44569);
    border: none;
    color: #ffffff;
    font-size: 15px;
    padding: 16px 32px;
    min-height: 30px;
}

QPushButton#generateBtn:hover {
    background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
        stop:0 #ff8fb3, stop:1 #e05580);
}

QPushButton#generateBtn:disabled {
    background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
        stop:0 #4a3a4e, stop:1 #3a2a3e);
    color: #7a7a8c;
}

QPushButton#exportBtn {
    background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
        stop:0 #6c5ce7, stop:1 #a29bfe);
    border: none;
    color: #ffffff;
}

QPushButton#exportBtn:hover {
    background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
        stop:0 #8075e8, stop:1 #b8b2ff);
}

QPushButton#modeBtn {
    padding: 8px 12px;
    font-size: 12px;
    border-radius: 8px;
    background: #1a1a2e;
    border: 2px solid #2d2d44;
    color: #e0e0e0;
    min-width: 130px;
    min-height: 70px;
}

QPushButton#modeBtn:hover {
    border-color: #00ffd5;
    color: #ffffff;
}

QPushButton#modeBtn:checked {
    background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
        stop:0 rgba(0, 255, 213, 0.2), stop:1 rgba(255, 107, 157, 0.2));
    border: 2px solid #00ffd5;
    color: #00ffd5;
}

QPushButton#clearBtn {
    background: transparent;
    border: 1px solid #ff6b6b;
    color: #ff6b6b;
    padding: 8px 16px;
    font-size: 12px;
    min-height: 16px;
    font-weight: 600;
}

QPushButton#clearBtn:hover {
    background: rgba(255, 107, 107, 0.2);
}

QPushButton#addImageBtn {
    background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
        stop:0 #1a1a2e, stop:1 #16213e);
    border: 1px solid #00ffd5;
    color: #00ffd5;
    padding: 8px 16px;
    font-size: 12px;
    min-height: 16px;
    font-weight: 600;
}

QPushButton#addImageBtn:hover {
    background: #00ffd5;
    color: #0a0a0f;
}

QPushButton#removeImageBtn {
    background: rgba(255, 107, 107, 0.8);
    border: none;
    color: white;
    padding: 2px 6px;
    font-size: 10px;
    font-weight: bold;
    border-radius: 10px;
    min-height: 16px;
    min-width: 16px;
    max-height: 20px;
    max-width: 20px;
}

QPushButton#removeImageBtn:hover {
    background: #ff6b6b;
}

QLineEdit {
    background: #1a1a2e;
    border: 1px solid #2d2d44;
    border-radius: 6px;
    padding: 10px 14px;
    font-size: 13px;
    color: #e0e0e0;
    selection-background-color: #00ffd5;
    selection-color: #0a0a0f;
}

QLineEdit:focus {
    border: 1px solid #00ffd5;
}

QTextEdit {
    background: #ffffff;
    border: 1px solid #2d2d44;
    border-radius: 8px;
    padding: 12px;
    font-size: 13px;
    color: #000000;
    line-height: 1.6;
    selection-background-color: #00ffd5;
    selection-color: #0a0a0f;
}

QTextEdit:focus {
    border: 1px solid #00ffd5;
}

QComboBox {
    background: #1a1a2e;
    border: 1px solid #2d2d44;
    border-radius: 6px;
    padding: 10px 14px;
    font-size: 13px;
    color: #e0e0e0;
    min-width: 150px;
}

QComboBox:hover {
    border: 1px solid #00ffd5;
}

QComboBox::drop-down {
    border: none;
    width: 30px;
}

QComboBox::down-arrow {
    image: none;
    border-left: 5px solid transparent;
    border-right: 5px solid transparent;
    border-top: 6px solid #00ffd5;
    margin-right: 10px;
}

QComboBox QAbstractItemView {
    background: #1a1a2e;
    border: 1px solid #00ffd5;
    selection-background-color: #00ffd5;
    selection-color: #0a0a0f;
    outline: none;
}

QGroupBox {
    font-size: 14px;
    font-weight: 600;
    color: #00ffd5;
    border: 1px solid #2d2d44;
    border-radius: 10px;
    margin-top: 16px;
    padding-top: 16px;
    background: rgba(26, 26, 46, 0.5);
}

QGroupBox::title {
    subcontrol-origin: margin;
    subcontrol-position: top left;
    left: 16px;
    padding: 0 8px;
    background: #12121a;
}

QCheckBox {
    font-size: 13px;
    color: #e0e0e0;
    spacing: 8px;
}

QCheckBox::indicator {
    width: 18px;
    height: 18px;
    border-radius: 4px;
    border: 1px solid #2d2d44;
    background: #1a1a2e;
}

QCheckBox::indicator:checked {
    background: #00ffd5;
    border: 1px solid #00ffd5;
}

QRadioButton {
    font-size: 13px;
    color: #e0e0e0;
    spacing: 8px;
}

QRadioButton::indicator {
    width: 18px;
    height: 18px;
    border-radius: 9px;
    border: 2px solid #2d2d44;
    background: #1a1a2e;
}

QRadioButton::indicator:checked {
    background: #00ffd5;
    border: 2px solid #00ffd5;
}

QProgressBar {
    background: #1a1a2e;
    border: none;
    border-radius: 6px;
    height: 8px;
    text-align: center;
}

QProgressBar::chunk {
    background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
        stop:0 #00ffd5, stop:1 #ff6b9d);
    border-radius: 6px;
}

QScrollArea {
    border: none;
    background: transparent;
}

QScrollBar:vertical {
    background: #12121a;
    width: 10px;
    border-radius: 5px;
    margin: 0;
}

QScrollBar::handle:vertical {
    background: #2d2d44;
    border-radius: 5px;
    min-height: 30px;
}

QScrollBar::handle:vertical:hover {
    background: #00ffd5;
}

QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical {
    height: 0;
}

QScrollBar:horizontal {
    background: #12121a;
    height: 10px;
    border-radius: 5px;
    margin: 0;
}

QScrollBar::handle:horizontal {
    background: #2d2d44;
    border-radius: 5px;
    min-width: 30px;
}

QScrollBar::handle:horizontal:hover {
    background: #00ffd5;
}

QScrollBar::add-line:horizontal, QScrollBar::sub-line:horizontal {
    width: 0;
}

QTabWidget::pane {
    border: 1px solid #2d2d44;
    border-radius: 8px;
    background: rgba(18, 18, 26, 0.8);
    top: -1px;
}

QTabBar::tab {
    background: #1a1a2e;
    border: 1px solid #2d2d44;
    border-bottom: none;
    border-top-left-radius: 8px;
    border-top-right-radius: 8px;
    padding: 10px 20px;
    margin-right: 4px;
    font-size: 13px;
    color: #7a7a8c;
}

QTabBar::tab:selected {
    background: #12121a;
    color: #00ffd5;
    border-color: #00ffd5;
    border-bottom: 1px solid #12121a;
}

QTabBar::tab:hover:!selected {
    color: #e0e0e0;
}

QFrame#dropZone {
    background: rgba(26, 26, 46, 0.6);
    border: 2px dashed #2d2d44;
    border-radius: 8px;
    min-height: 60px;
    max-height: 70px;
}

QFrame#dropZone:hover {
    border-color: #00ffd5;
    background: rgba(0, 255, 213, 0.05);
}

QFrame#previewFrame {
    background: #12121a;
    border: 1px solid #2d2d44;
    border-radius: 8px;
    padding: 8px;
}

QFrame#resultCard {
    background: rgba(26, 26, 46, 0.7);
    border: 1px solid #2d2d44;
    border-radius: 10px;
    padding: 16px;
    margin: 8px 0;
}

QFrame#resultCard:hover {
    border-color: #00ffd5;
}

QFrame#modeSelectorFrame {
    background: rgba(26, 26, 46, 0.8);
    border: 1px solid #2d2d44;
    border-radius: 12px;
    padding: 16px;
}

QFrame#inputStatusFrame {
    background: rgba(18, 18, 26, 0.9);
    border: 1px solid #2d2d44;
    border-radius: 8px;
    padding: 12px;
}

QFrame#imageThumbnail {
    background: #12121a;
    border: 1px solid #2d2d44;
    border-radius: 8px;
    padding: 4px;
}

QFrame#imageThumbnail:hover {
    border-color: #00ffd5;
}
"""


# ═══════════════════════════════════════════════════════════════════════════════
# GEMINI API WORKER THREAD
# ═══════════════════════════════════════════════════════════════════════════════

class GeminiWorker(QThread):
    """Background worker for Gemini API calls."""
    
    finished = pyqtSignal(str)
    error = pyqtSignal(str)
    progress = pyqtSignal(int)
    
    def __init__(self, api_key: str, image_paths: List[str], document_content: Optional[str],
                 test_type: str, app_context: str, input_mode: str):
        super().__init__()
        self.api_key = api_key
        self.image_paths = image_paths  # Now a list of paths
        self.document_content = document_content
        self.test_type = test_type
        self.app_context = app_context
        self.input_mode = input_mode
    
    def run(self):
        try:
            self.progress.emit(10)
            genai.configure(api_key=self.api_key)
            
            self.progress.emit(20)
            
            # Use gemini-1.5-flash for all cases
            model = genai.GenerativeModel('gemini-2.5-flash')
            
            self.progress.emit(40)
            
            # Build the prompt
            prompt = self._build_prompt()
            
            self.progress.emit(50)
            
            # Prepare content for the API call
            content = []
            
            # Add all images if we're in image or combined mode
            if self.image_paths and self.input_mode in ("image", "combined"):
                for img_path in self.image_paths:
                    img = Image.open(img_path)
                    content.append(img)
            
            content.append(prompt)
            
            self.progress.emit(70)
            
            # Generate response
            response = model.generate_content(content)
            
            self.progress.emit(90)
            
            result = response.text
            
            self.progress.emit(100)
            self.finished.emit(result)
            
        except Exception as e:
            self.error.emit(str(e))
    
    def _build_prompt(self) -> str:
        """Build the prompt for test case generation."""
        
        num_images = len(self.image_paths) if self.image_paths else 0
        
        base_prompt = f"""You are an expert Android QA Engineer and Test Automation Specialist. 
Your task is to generate comprehensive test cases for an Android mobile application.

**Test Type Requested:** {self.test_type}

**Application Context:** {self.app_context if self.app_context else "General Android application"}

"""
        
        if self.input_mode == "combined" and self.image_paths and self.document_content:
            base_prompt += f"""**Input Mode:** COMBINED (Images + Document)

You have been provided with:
1. {num_images} screenshot(s)/image(s) of the Android app screens
2. A document with additional specifications/requirements

Analyze ALL images and the document content together to generate comprehensive test cases. The images show the visual UI and user flow while the document provides context, requirements, and specifications.

**Document Content:**
""" + self.document_content + """

"""
        elif self.input_mode == "image" and self.image_paths:
            base_prompt += f"""**Input Mode:** IMAGE ONLY ({num_images} image(s) provided)

Analyze the provided {num_images} screenshot(s)/image(s) to identify:
1. UI elements (buttons, text fields, lists, dropdowns, etc.)
2. Navigation patterns and screen flow indicators
3. User interaction points and touch targets
4. Visual states and feedback indicators
5. Relationships between different screens (if multiple images)
6. Potential edge cases based on visible UI elements

"""
        elif self.input_mode == "document" and self.document_content:
            base_prompt += """**Input Mode:** DOCUMENT ONLY

Analyze the provided document containing specifications/requirements:

**Document Content:**
""" + self.document_content + """

"""
        
        base_prompt += """
**Generate Test Cases in this CLEAN, EASY-TO-READ format:**

══════════════════════════════════════════════════════════════════════════════
                         TEST SUITE: [Feature Name]
══════════════════════════════════════════════════════════════════════════════

For EACH test case, use this EXACT format:

──────────────────────────────────────────────────────────────────────────────
TEST CASE ID:     TC-001
TEST CASE NAME:   [Clear descriptive title]
──────────────────────────────────────────────────────────────────────────────

    PRIORITY:       [High / Medium / Low]
    TYPE:           [Functional / UI / Security / Performance]
    
    ▸ PRECONDITIONS:
      - Condition 1
      - Condition 2
    
    ▸ TEST STEPS:
      Step 1:  [First action to perform]
      Step 2:  [Second action to perform]
      Step 3:  [Third action to perform]
    
    ▸ EXPECTED RESULTS:
      ✓ Result 1
      ✓ Result 2
      ✓ Result 3
    
    ▸ TEST DATA:
      [Any input data, test values, or N/A]
    
    ▸ AUTOMATION NOTES:
      [Espresso/Appium hints, element IDs, locators]

══════════════════════════════════════════════════════════════════════════════


FORMATTING RULES:
1. Use ══════ for major section dividers
2. Use ────── for test case separators
3. Use ▸ for section headers within test case
4. Use ✓ for expected results
5. Use - for list items
6. Use Step 1:, Step 2: etc for numbered steps
7. Keep each section clearly labeled and indented
8. Add blank lines between sections for readability

Generate 10-15 comprehensive test cases covering:
- Positive scenarios (happy path)
- Negative scenarios (error handling, invalid inputs)
- Boundary conditions (min/max values, empty states)
- UI/UX validation (layout, responsiveness)
- Edge cases (network issues, interruptions)
- Security considerations (input validation)

Include Android-specific considerations:
- Back button behavior
- App lifecycle (background/foreground)
- Different screen sizes
- Accessibility (TalkBack, font scaling)
"""
        
        return base_prompt


# ═══════════════════════════════════════════════════════════════════════════════
# DROP ZONE WIDGET
# ═══════════════════════════════════════════════════════════════════════════════

class DropZone(QFrame):
    """Custom drop zone for files."""
    
    fileDropped = pyqtSignal(str)
    
    def __init__(self, file_type: str, parent=None):
        super().__init__(parent)
        self.file_type = file_type
        self.setAcceptDrops(True)
        self.setObjectName("dropZone")
        self.setup_ui()
    
    def setup_ui(self):
        layout = QHBoxLayout(self)
        layout.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout.setSpacing(12)
        layout.setContentsMargins(12, 6, 12, 6)
        
        # Icon label (using emoji as placeholder)
        icon_label = QLabel("📷" if self.file_type == "image" else "📄")
        icon_label.setStyleSheet("font-size: 24px;")
        icon_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        
        # Text container
        text_layout = QVBoxLayout()
        text_layout.setSpacing(2)
        
        # Main text
        self.main_label = QLabel(
            f"Drop {'Image(s)' if self.file_type == 'image' else 'Document'} Here"
        )
        self.main_label.setStyleSheet("""
            font-size: 12px;
            font-weight: 600;
            color: #e0e0e0;
        """)
        
        # Sub text
        formats = "PNG, JPG, JPEG, GIF" if self.file_type == "image" else "PDF, DOCX, TXT"
        self.sub_label = QLabel(f"Click to browse • {formats}")
        self.sub_label.setStyleSheet("""
            font-size: 9px;
            color: #7a7a8c;
        """)
        
        # File name display
        self.file_label = QLabel("")
        self.file_label.setStyleSheet("""
            font-size: 10px;
            color: #00ffd5;
            font-weight: 600;
        """)
        self.file_label.hide()
        
        text_layout.addWidget(self.main_label)
        text_layout.addWidget(self.sub_label)
        text_layout.addWidget(self.file_label)
        
        layout.addWidget(icon_label)
        layout.addLayout(text_layout)
        layout.addStretch()
    
    def dragEnterEvent(self, event: QDragEnterEvent):
        if event.mimeData().hasUrls():
            event.acceptProposedAction()
            self.setStyleSheet("""
                QFrame#dropZone {
                    border-color: #00ffd5;
                    background: rgba(0, 255, 213, 0.1);
                }
            """)
    
    def dragLeaveEvent(self, event):
        self.setStyleSheet("")
    
    def dropEvent(self, event: QDropEvent):
        self.setStyleSheet("")
        files = [url.toLocalFile() for url in event.mimeData().urls()]
        for f in files:
            self.fileDropped.emit(f)
    
    def mousePressEvent(self, event):
        if self.file_type == "image":
            file_filter = "Images (*.png *.jpg *.jpeg *.gif *.bmp)"
            files, _ = QFileDialog.getOpenFileNames(
                self, "Select Images", "", file_filter
            )
            for f in files:
                self.fileDropped.emit(f)
        else:
            file_filter = "Documents (*.pdf *.docx *.txt *.md)"
            file_path, _ = QFileDialog.getOpenFileName(
                self, "Select Document", "", file_filter
            )
            if file_path:
                self.fileDropped.emit(file_path)
    
    def set_file(self, path: str):
        """Update UI to show selected file."""
        filename = Path(path).name
        self.file_label.setText(f"✓ {filename}")
        self.file_label.show()
        self.main_label.setText("File Selected")
        self.sub_label.setText("Click to change")
    
    def reset(self):
        """Reset to initial state."""
        formats = "PNG, JPG, JPEG, GIF" if self.file_type == "image" else "PDF, DOCX, TXT"
        self.main_label.setText(f"Drop {'Image' if self.file_type == 'image' else 'Document'} Here")
        self.sub_label.setText(f"or click to browse • Supports {formats}")
        self.file_label.hide()


# ═══════════════════════════════════════════════════════════════════════════════
# IMAGE THUMBNAIL WIDGET
# ═══════════════════════════════════════════════════════════════════════════════

class ImageThumbnail(QFrame):
    """Widget to display an image thumbnail with remove button."""
    
    removeClicked = pyqtSignal(str)  # Emits the image path when remove is clicked
    
    def __init__(self, image_path: str, parent=None):
        super().__init__(parent)
        self.image_path = image_path
        self.setObjectName("imageThumbnail")
        self.setFixedSize(75, 90)
        self.setup_ui()
    
    def setup_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(2, 2, 2, 2)
        layout.setSpacing(2)
        
        # Container for image and remove button
        img_container = QWidget()
        img_container.setFixedSize(70, 60)
        
        # Image label
        self.image_label = QLabel(img_container)
        self.image_label.setFixedSize(70, 60)
        self.image_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.image_label.setStyleSheet("""
            background: #0a0a0f;
            border-radius: 4px;
        """)
        
        # Load and scale image
        pixmap = QPixmap(self.image_path)
        scaled = pixmap.scaled(
            66, 56,
            Qt.AspectRatioMode.KeepAspectRatio,
            Qt.TransformationMode.SmoothTransformation
        )
        self.image_label.setPixmap(scaled)
        
        # Remove button (positioned at top-right)
        self.remove_btn = QPushButton("×", img_container)
        self.remove_btn.setObjectName("removeImageBtn")
        self.remove_btn.setFixedSize(18, 18)
        self.remove_btn.move(50, 2)
        self.remove_btn.clicked.connect(lambda: self.removeClicked.emit(self.image_path))
        self.remove_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        
        # Filename label
        filename = Path(self.image_path).name
        name_label = QLabel(filename[:10] + ".." if len(filename) > 10 else filename)
        name_label.setStyleSheet("font-size: 8px; color: #7a7a8c;")
        name_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        
        layout.addWidget(img_container)
        layout.addWidget(name_label)


# ═══════════════════════════════════════════════════════════════════════════════
# MULTI-IMAGE PREVIEW WIDGET
# ═══════════════════════════════════════════════════════════════════════════════

class MultiImagePreview(QFrame):
    """Widget to display multiple image thumbnails in a scrollable area."""
    
    imageRemoved = pyqtSignal(str)
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self.thumbnails = {}
        self.setFixedHeight(90)
        self.setStyleSheet("""
            QFrame {
                background: rgba(18, 18, 26, 0.6);
                border: 1px solid #2d2d44;
                border-radius: 6px;
            }
        """)
        self.setup_ui()
    
    def setup_ui(self):
        main_layout = QHBoxLayout(self)
        main_layout.setContentsMargins(8, 8, 8, 8)
        main_layout.setSpacing(8)
        
        # Header with count
        self.count_label = QLabel("📷\n0")
        self.count_label.setStyleSheet("font-size: 12px; color: #00ffd5; font-weight: 700;")
        self.count_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.count_label.setFixedWidth(35)
        main_layout.addWidget(self.count_label)
        
        # Scroll area for thumbnails
        self.scroll = QScrollArea()
        self.scroll.setWidgetResizable(True)
        self.scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAsNeeded)
        self.scroll.setVerticalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        self.scroll.setStyleSheet("""
            QScrollArea {
                background: transparent;
                border: none;
            }
        """)
        
        # Container widget
        self.container = QWidget()
        self.container_layout = QHBoxLayout(self.container)
        self.container_layout.setContentsMargins(0, 0, 0, 0)
        self.container_layout.setSpacing(6)
        self.container_layout.setAlignment(Qt.AlignmentFlag.AlignLeft)
        
        # Placeholder label
        self.placeholder = QLabel("No images yet - drop or click above")
        self.placeholder.setStyleSheet("color: #aaaacc; font-size: 12px;")
        self.placeholder.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.container_layout.addWidget(self.placeholder)
        
        self.scroll.setWidget(self.container)
        main_layout.addWidget(self.scroll)
    
    def add_image(self, image_path: str):
        """Add an image thumbnail."""
        if image_path in self.thumbnails:
            return  # Already exists
        
        # Hide placeholder if this is the first image
        if self.placeholder.isVisible():
            self.placeholder.hide()
        
        # Create thumbnail
        thumbnail = ImageThumbnail(image_path)
        thumbnail.removeClicked.connect(self._on_remove_clicked)
        
        self.thumbnails[image_path] = thumbnail
        self.container_layout.addWidget(thumbnail)
        self._update_count()
    
    def remove_image(self, image_path: str):
        """Remove an image thumbnail."""
        if image_path in self.thumbnails:
            thumbnail = self.thumbnails.pop(image_path)
            self.container_layout.removeWidget(thumbnail)
            thumbnail.deleteLater()
            self._update_count()
            
            # Show placeholder if no images left
            if not self.thumbnails:
                self.placeholder.show()
    
    def _on_remove_clicked(self, image_path: str):
        """Handle remove button click."""
        self.remove_image(image_path)
        self.imageRemoved.emit(image_path)
    
    def _update_count(self):
        """Update the image count label."""
        count = len(self.thumbnails)
        self.count_label.setText(f"📷\n{count}")
    
    def clear(self):
        """Remove all thumbnails."""
        for path in list(self.thumbnails.keys()):
            self.remove_image(path)
        self.placeholder.show()
    
    def get_image_paths(self) -> List[str]:
        """Get all image paths."""
        return list(self.thumbnails.keys())


# ═══════════════════════════════════════════════════════════════════════════════
# INPUT MODE BUTTON
# ═══════════════════════════════════════════════════════════════════════════════

class ModeButton(QPushButton):
    """Checkable button for input mode selection."""
    
    def __init__(self, text: str, icon_emoji: str, description: str, parent=None):
        super().__init__(parent)
        self.setCheckable(True)
        self.setObjectName("modeBtn")
        self.setCursor(Qt.CursorShape.PointingHandCursor)
        self.setMinimumWidth(140)
        self.setMinimumHeight(80)
        
        # Create layout
        layout = QVBoxLayout(self)
        layout.setSpacing(2)
        layout.setContentsMargins(8, 8, 8, 8)
        
        # Icon
        icon = QLabel(icon_emoji)
        icon.setStyleSheet("font-size: 20px; background: transparent;")
        icon.setAlignment(Qt.AlignmentFlag.AlignCenter)
        icon.setAttribute(Qt.WidgetAttribute.WA_TransparentForMouseEvents)
        
        # Text
        title = QLabel(text)
        title.setStyleSheet("font-size: 13px; font-weight: 700; background: transparent; color: #ffffff;")
        title.setAlignment(Qt.AlignmentFlag.AlignCenter)
        title.setAttribute(Qt.WidgetAttribute.WA_TransparentForMouseEvents)
        
        # Description
        desc = QLabel(description)
        desc.setStyleSheet("font-size: 10px; color: #b0b0c0; background: transparent;")
        desc.setAlignment(Qt.AlignmentFlag.AlignCenter)
        desc.setAttribute(Qt.WidgetAttribute.WA_TransparentForMouseEvents)
        
        layout.addWidget(icon)
        layout.addWidget(title)
        layout.addWidget(desc)


# ═══════════════════════════════════════════════════════════════════════════════
# MAIN WINDOW
# ═══════════════════════════════════════════════════════════════════════════════

class AndroidTestCaseGenerator(QMainWindow):
    """Main application window."""
    
    def __init__(self):
        super().__init__()
        self.image_paths: List[str] = []  # Changed to list
        self.document_path: Optional[str] = None
        self.document_content: Optional[str] = None
        self.worker: Optional[GeminiWorker] = None
        self.input_mode = "image"  # "image", "document", "combined"
        
        self.setup_ui()
        self.setWindowTitle("Android Test Case Generator")
        self.setMinimumSize(1200, 800)
        self.resize(1400, 900)
        self._update_mode_ui()
    
    def setup_ui(self):
        """Setup the main UI."""
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        
        main_layout = QVBoxLayout(central_widget)
        main_layout.setSpacing(0)
        main_layout.setContentsMargins(0, 0, 0, 0)
        
        # Header
        header = self._create_header()
        main_layout.addWidget(header)
        
        # Content area with splitter
        content_splitter = QSplitter(Qt.Orientation.Horizontal)
        content_splitter.setContentsMargins(24, 24, 24, 24)
        
        # Left panel - Input
        left_panel = self._create_input_panel()
        content_splitter.addWidget(left_panel)
        
        # Right panel - Output
        right_panel = self._create_output_panel()
        content_splitter.addWidget(right_panel)
        
        content_splitter.setSizes([550, 650])
        content_splitter.setHandleWidth(1)
        
        main_layout.addWidget(content_splitter, 1)
        
        # Footer
        footer = self._create_footer()
        main_layout.addWidget(footer)
    
    def _create_header(self) -> QWidget:
        """Create the header section."""
        header = QFrame()
        header.setStyleSheet("""
            QFrame {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 rgba(0, 255, 213, 0.1), stop:1 rgba(255, 107, 157, 0.1));
                border-bottom: 1px solid #2d2d44;
            }
        """)
        
        layout = QHBoxLayout(header)
        layout.setContentsMargins(32, 20, 32, 20)
        
        # Left side - Title
        title_layout = QVBoxLayout()
        title_layout.setSpacing(4)
        
        title = QLabel("ANDROID TEST GENERATOR")
        title.setObjectName("title")
        
        subtitle = QLabel("Generate comprehensive test cases from images or documents using Gemini AI")
        subtitle.setObjectName("subtitle")
        
        title_layout.addWidget(title)
        title_layout.addWidget(subtitle)
        
        # Right side - API Key
        api_layout = QHBoxLayout()
        api_layout.setSpacing(12)
        
        api_label = QLabel("Gemini API Key:")
        api_label.setStyleSheet("color: #7a7a8c;")
        
        self.api_key_input = QLineEdit()
        self.api_key_input.setPlaceholderText("Enter your Gemini API key...")
        self.api_key_input.setEchoMode(QLineEdit.EchoMode.Password)
        self.api_key_input.setFixedWidth(300)
        
        # Check for env variable or use default key
        env_key = os.environ.get("GEMINI_API_KEY", "")
        if env_key:
            self.api_key_input.setText(env_key)
        else:
            # Default API key
            self.api_key_input.setText("AIzaSyCjX4PQrVTEg0W17YxV_2ikF11b2jCMs5w")
        
        api_layout.addWidget(api_label)
        api_layout.addWidget(self.api_key_input)
        
        layout.addLayout(title_layout)
        layout.addStretch()
        layout.addLayout(api_layout)
        
        return header
    
    def _create_input_panel(self) -> QWidget:
        """Create the input panel."""
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        
        panel = QFrame()
        panel.setStyleSheet("QFrame { background: transparent; }")
        
        layout = QVBoxLayout(panel)
        layout.setSpacing(16)
        layout.setContentsMargins(12, 12, 12, 12)
        
        # Section title
        input_title = QLabel("INPUT CONFIGURATION")
        input_title.setObjectName("sectionTitle")
        layout.addWidget(input_title)
        
        # ═══════════════════════════════════════════════════════════════════
        # MODE SELECTOR
        # ═══════════════════════════════════════════════════════════════════
        mode_frame = QFrame()
        mode_frame.setObjectName("modeSelectorFrame")
        mode_layout = QVBoxLayout(mode_frame)
        mode_layout.setSpacing(12)
        
        mode_title = QLabel("Select Input Source")
        mode_title.setObjectName("modeTitle")
        mode_layout.addWidget(mode_title)
        
        # Mode buttons in horizontal layout
        buttons_layout = QHBoxLayout()
        buttons_layout.setSpacing(12)
        
        self.mode_group = QButtonGroup(self)
        self.mode_group.setExclusive(True)
        
        self.image_mode_btn = ModeButton("Image Only", "📷", "Screenshot analysis")
        self.doc_mode_btn = ModeButton("Document Only", "📄", "From specifications")
        self.combined_mode_btn = ModeButton("Combined", "🔗", "Image + Document")
        
        self.mode_group.addButton(self.image_mode_btn, 0)
        self.mode_group.addButton(self.doc_mode_btn, 1)
        self.mode_group.addButton(self.combined_mode_btn, 2)
        
        self.image_mode_btn.setChecked(True)
        
        self.mode_group.buttonClicked.connect(self._on_mode_changed)
        
        buttons_layout.addWidget(self.image_mode_btn)
        buttons_layout.addWidget(self.doc_mode_btn)
        buttons_layout.addWidget(self.combined_mode_btn)
        
        mode_layout.addLayout(buttons_layout)
        layout.addWidget(mode_frame)
        
        # ═══════════════════════════════════════════════════════════════════
        # INPUT STATUS INDICATOR
        # ═══════════════════════════════════════════════════════════════════
        status_frame = QFrame()
        status_frame.setObjectName("inputStatusFrame")
        status_layout = QHBoxLayout(status_frame)
        status_layout.setSpacing(16)
        
        self.image_status = QLabel("📷 No images")
        self.image_status.setObjectName("statusInactive")
        
        self.doc_status = QLabel("📄 No document")
        self.doc_status.setObjectName("statusInactive")
        
        status_layout.addWidget(self.image_status)
        status_layout.addWidget(self.doc_status)
        status_layout.addStretch()
        
        layout.addWidget(status_frame)
        
        # ═══════════════════════════════════════════════════════════════════
        # STACKED WIDGET FOR INPUT AREAS
        # ═══════════════════════════════════════════════════════════════════
        self.input_stack = QStackedWidget()
        
        # Page 0: Image input
        image_page = self._create_image_input_page()
        self.input_stack.addWidget(image_page)
        
        # Page 1: Document input
        doc_page = self._create_document_input_page()
        self.input_stack.addWidget(doc_page)
        
        # Page 2: Combined input
        combined_page = self._create_combined_input_page()
        self.input_stack.addWidget(combined_page)
        
        layout.addWidget(self.input_stack)
        
        # ═══════════════════════════════════════════════════════════════════
        # OPTIONS GROUP
        # ═══════════════════════════════════════════════════════════════════
        options_group = QGroupBox("Generation Options")
        options_layout = QGridLayout(options_group)
        options_layout.setSpacing(12)
        
        # Test type
        type_label = QLabel("Test Type:")
        self.test_type_combo = QComboBox()
        self.test_type_combo.addItems([
            "Comprehensive (All Types)",
            "Functional Testing",
            "UI/UX Testing",
            "Integration Testing",
            "Performance Testing",
            "Security Testing",
            "Accessibility Testing"
        ])
        options_layout.addWidget(type_label, 0, 0)
        options_layout.addWidget(self.test_type_combo, 0, 1)
        
        # App context
        context_label = QLabel("App Context:")
        self.context_input = QLineEdit()
        self.context_input.setPlaceholderText("e.g., E-commerce app, Login screen...")
        options_layout.addWidget(context_label, 1, 0)
        options_layout.addWidget(self.context_input, 1, 1)
        
        # Checkboxes
        self.include_automation = QCheckBox("Include automation hints")
        self.include_automation.setChecked(True)
        self.include_data = QCheckBox("Include test data suggestions")
        self.include_data.setChecked(True)
        
        options_layout.addWidget(self.include_automation, 2, 0)
        options_layout.addWidget(self.include_data, 2, 1)
        
        layout.addWidget(options_group)
        
        # ═══════════════════════════════════════════════════════════════════
        # ACTION BUTTONS
        # ═══════════════════════════════════════════════════════════════════
        actions_layout = QHBoxLayout()
        
        self.clear_all_btn = QPushButton("Clear All")
        self.clear_all_btn.setObjectName("clearBtn")
        self.clear_all_btn.clicked.connect(self._clear_all)
        actions_layout.addWidget(self.clear_all_btn)
        
        actions_layout.addStretch()
        layout.addLayout(actions_layout)
        
        # Generate button
        self.generate_btn = QPushButton("⚡ GENERATE TEST CASES")
        self.generate_btn.setObjectName("generateBtn")
        self.generate_btn.clicked.connect(self._generate_test_cases)
        layout.addWidget(self.generate_btn)
        
        # Progress bar
        self.progress_bar = QProgressBar()
        self.progress_bar.setVisible(False)
        self.progress_bar.setTextVisible(False)
        layout.addWidget(self.progress_bar)
        
        layout.addStretch()
        
        scroll.setWidget(panel)
        return scroll
    
    def _create_image_input_page(self) -> QWidget:
        """Create the image-only input page."""
        page = QWidget()
        layout = QVBoxLayout(page)
        layout.setSpacing(4)
        layout.setContentsMargins(0, 0, 0, 0)
        
        # Drop zone
        self.image_drop = DropZone("image")
        self.image_drop.fileDropped.connect(self._on_image_dropped)
        layout.addWidget(self.image_drop)
        
        # Multi-image preview
        self.image_preview = MultiImagePreview()
        self.image_preview.imageRemoved.connect(self._on_image_removed)
        layout.addWidget(self.image_preview)
        
        # Buttons row
        btn_row = QHBoxLayout()
        btn_row.setSpacing(8)
        
        add_more_btn = QPushButton("+ Add More")
        add_more_btn.setObjectName("addImageBtn")
        add_more_btn.clicked.connect(self._browse_more_images)
        btn_row.addWidget(add_more_btn)
        
        clear_btn = QPushButton("Clear All")
        clear_btn.setObjectName("clearBtn")
        clear_btn.clicked.connect(self._clear_images)
        btn_row.addWidget(clear_btn)
        
        btn_row.addStretch()
        layout.addLayout(btn_row)
        
        return page
    
    def _create_document_input_page(self) -> QWidget:
        """Create the document-only input page."""
        page = QWidget()
        layout = QVBoxLayout(page)
        layout.setSpacing(12)
        layout.setContentsMargins(0, 0, 0, 0)
        
        self.doc_drop = DropZone("document")
        self.doc_drop.fileDropped.connect(self._on_document_dropped)
        layout.addWidget(self.doc_drop)
        
        # Preview
        preview_group = QGroupBox("Document Content Preview")
        preview_layout = QVBoxLayout(preview_group)
        self.doc_preview = QTextEdit()
        self.doc_preview.setReadOnly(True)
        self.doc_preview.setPlaceholderText("Document content will appear here...")
        self.doc_preview.setMaximumHeight(150)
        preview_layout.addWidget(self.doc_preview)
        
        clear_btn = QPushButton("Clear Document")
        clear_btn.setObjectName("clearBtn")
        clear_btn.clicked.connect(self._clear_document)
        preview_layout.addWidget(clear_btn)
        
        layout.addWidget(preview_group)
        
        return page
    
    def _create_combined_input_page(self) -> QWidget:
        """Create the combined input page."""
        page = QWidget()
        layout = QVBoxLayout(page)
        layout.setSpacing(12)
        layout.setContentsMargins(0, 0, 0, 0)
        
        # Images section
        image_group = QGroupBox("📷 Screenshots")
        image_layout = QVBoxLayout(image_group)
        
        self.combined_image_drop = DropZone("image")
        self.combined_image_drop.fileDropped.connect(self._on_image_dropped)
        image_layout.addWidget(self.combined_image_drop)
        
        self.combined_image_preview = MultiImagePreview()
        self.combined_image_preview.imageRemoved.connect(self._on_image_removed)
        image_layout.addWidget(self.combined_image_preview)
        
        # Buttons for images
        img_btn_layout = QHBoxLayout()
        add_img_btn = QPushButton("➕ Add Images")
        add_img_btn.setObjectName("addImageBtn")
        add_img_btn.clicked.connect(self._browse_more_images)
        img_btn_layout.addWidget(add_img_btn)
        img_btn_layout.addStretch()
        image_layout.addLayout(img_btn_layout)
        
        layout.addWidget(image_group)
        
        # Document section
        doc_group = QGroupBox("📄 Document")
        doc_layout = QVBoxLayout(doc_group)
        
        self.combined_doc_drop = DropZone("document")
        self.combined_doc_drop.fileDropped.connect(self._on_document_dropped)
        doc_layout.addWidget(self.combined_doc_drop)
        
        self.combined_doc_preview = QTextEdit()
        self.combined_doc_preview.setReadOnly(True)
        self.combined_doc_preview.setPlaceholderText("Document content...")
        self.combined_doc_preview.setMaximumHeight(100)
        doc_layout.addWidget(self.combined_doc_preview)
        
        layout.addWidget(doc_group)
        
        # Clear buttons
        clear_layout = QHBoxLayout()
        clear_img_btn = QPushButton("Clear Images")
        clear_img_btn.setObjectName("clearBtn")
        clear_img_btn.clicked.connect(self._clear_images)
        
        clear_doc_btn = QPushButton("Clear Document")
        clear_doc_btn.setObjectName("clearBtn")
        clear_doc_btn.clicked.connect(self._clear_document)
        
        clear_layout.addWidget(clear_img_btn)
        clear_layout.addWidget(clear_doc_btn)
        clear_layout.addStretch()
        
        layout.addLayout(clear_layout)
        
        return page
    
    def _create_output_panel(self) -> QWidget:
        """Create the output panel."""
        panel = QFrame()
        panel.setStyleSheet("QFrame { background: transparent; }")
        
        layout = QVBoxLayout(panel)
        layout.setSpacing(16)
        layout.setContentsMargins(12, 12, 12, 12)
        
        # Header with title and export button
        header_layout = QHBoxLayout()
        
        output_title = QLabel("GENERATED TEST CASES")
        output_title.setObjectName("sectionTitle")
        header_layout.addWidget(output_title)
        
        header_layout.addStretch()
        
        self.copy_btn = QPushButton("📋 Copy")
        self.copy_btn.clicked.connect(self._copy_results)
        self.copy_btn.setEnabled(False)
        header_layout.addWidget(self.copy_btn)
        
        self.export_btn = QPushButton("📥 Export")
        self.export_btn.setObjectName("exportBtn")
        self.export_btn.clicked.connect(self._export_results)
        self.export_btn.setEnabled(False)
        header_layout.addWidget(self.export_btn)
        
        layout.addLayout(header_layout)
        
        # Results area
        self.results_text = QTextEdit()
        self.results_text.setReadOnly(True)
        self.results_text.setPlaceholderText(
            "Generated test cases will appear here...\n\n"
            "━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n\n"
            "HOW TO USE:\n\n"
            "1️⃣  Enter your Gemini API key (top right)\n\n"
            "2️⃣  Choose input mode:\n"
            "     • Image Only - Upload screenshot(s)\n"
            "     • Document Only - Upload specs/requirements\n"
            "     • Combined - Use both for best results\n\n"
            "3️⃣  Upload your file(s) - Multiple images supported!\n\n"
            "4️⃣  (Optional) Set test type and app context\n\n"
            "5️⃣  Click 'Generate Test Cases'\n\n"
            "━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━"
        )
        layout.addWidget(self.results_text, 1)
        
        # Stats footer
        self.stats_label = QLabel("")
        self.stats_label.setStyleSheet("color: #7a7a8c; font-size: 12px;")
        layout.addWidget(self.stats_label)
        
        return panel
    
    def _create_footer(self) -> QWidget:
        """Create the footer section."""
        footer = QFrame()
        footer.setStyleSheet("""
            QFrame {
                background: #0a0a0f;
                border-top: 1px solid #2d2d44;
            }
        """)
        
        layout = QHBoxLayout(footer)
        layout.setContentsMargins(32, 12, 32, 12)
        
        self.status_label = QLabel("Ready • Select input mode and upload files")
        self.status_label.setStyleSheet("color: #7a7a8c; font-size: 11px;")
        
        version = QLabel("v2.1.0 • Powered by Google Gemini AI")
        version.setStyleSheet("color: #7a7a8c; font-size: 11px;")
        
        layout.addWidget(self.status_label)
        layout.addStretch()
        layout.addWidget(version)
        
        return footer
    
    # ═══════════════════════════════════════════════════════════════════════════
    # MODE HANDLING
    # ═══════════════════════════════════════════════════════════════════════════
    
    def _on_mode_changed(self, button):
        """Handle input mode change."""
        if button == self.image_mode_btn:
            self.input_mode = "image"
            self.input_stack.setCurrentIndex(0)
        elif button == self.doc_mode_btn:
            self.input_mode = "document"
            self.input_stack.setCurrentIndex(1)
        elif button == self.combined_mode_btn:
            self.input_mode = "combined"
            self.input_stack.setCurrentIndex(2)
        
        self._update_mode_ui()
    
    def _update_mode_ui(self):
        """Update UI based on current mode and inputs."""
        # Update status labels
        num_images = len(self.image_paths)
        if num_images > 0:
            self.image_status.setText(f"📷 {num_images} image(s)")
            self.image_status.setObjectName("statusActive")
        else:
            self.image_status.setText("📷 No images")
            self.image_status.setObjectName("statusInactive")
        
        if self.document_content:
            filename = Path(self.document_path).name if self.document_path else "Document"
            self.doc_status.setText(f"📄 {filename[:20]}..." if len(filename) > 20 else f"📄 {filename}")
            self.doc_status.setObjectName("statusActive")
        else:
            self.doc_status.setText("📄 No document")
            self.doc_status.setObjectName("statusInactive")
        
        # Force style refresh
        self.image_status.setStyle(self.image_status.style())
        self.doc_status.setStyle(self.doc_status.style())
        
        # Update status bar
        mode_text = {
            "image": "Image Only mode",
            "document": "Document Only mode",
            "combined": "Combined mode (Image + Document)"
        }
        self.status_label.setText(f"Ready • {mode_text[self.input_mode]}")
    
    # ═══════════════════════════════════════════════════════════════════════════
    # EVENT HANDLERS
    # ═══════════════════════════════════════════════════════════════════════════
    
    def _on_image_dropped(self, path: str):
        """Handle image file drop."""
        valid_extensions = {'.png', '.jpg', '.jpeg', '.gif', '.bmp'}
        ext = Path(path).suffix.lower()
        
        if ext not in valid_extensions:
            QMessageBox.warning(self, "Invalid File", "Please select a valid image file (PNG, JPG, JPEG, GIF, BMP).")
            return
        
        if path not in self.image_paths:
            self.image_paths.append(path)
            
            # Update all preview widgets
            self.image_preview.add_image(path)
            self.combined_image_preview.add_image(path)
        
        self._update_mode_ui()
    
    def _on_image_removed(self, path: str):
        """Handle image removal from preview."""
        if path in self.image_paths:
            self.image_paths.remove(path)
            
            # Sync both preview widgets
            self.image_preview.remove_image(path)
            self.combined_image_preview.remove_image(path)
        
        self._update_mode_ui()
    
    def _browse_more_images(self):
        """Open file dialog to add more images."""
        file_filter = "Images (*.png *.jpg *.jpeg *.gif *.bmp)"
        files, _ = QFileDialog.getOpenFileNames(
            self, "Select Images", "", file_filter
        )
        for f in files:
            self._on_image_dropped(f)
    
    def _on_document_dropped(self, path: str):
        """Handle document file drop."""
        ext = Path(path).suffix.lower()
        
        try:
            if ext == '.pdf':
                content = self._read_pdf(path)
            elif ext == '.docx':
                content = self._read_docx(path)
            elif ext in {'.txt', '.md'}:
                with open(path, 'r', encoding='utf-8') as f:
                    content = f.read()
            else:
                QMessageBox.warning(self, "Invalid File", "Please select a PDF, DOCX, TXT, or MD file.")
                return
            
            self.document_path = path
            self.document_content = content
            
            # Update all relevant UI elements
            preview_text = content[:2000] + "..." if len(content) > 2000 else content
            
            self.doc_drop.set_file(path)
            self.doc_preview.setText(preview_text)
            self.combined_doc_drop.set_file(path)
            self.combined_doc_preview.setText(preview_text)
            
            self._update_mode_ui()
            
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to read document: {str(e)}")
    
    def _read_pdf(self, path: str) -> str:
        """Extract text from PDF."""
        text = []
        with open(path, 'rb') as f:
            reader = PyPDF2.PdfReader(f)
            for page in reader.pages:
                page_text = page.extract_text()
                if page_text:
                    text.append(page_text)
        return '\n'.join(text)
    
    def _read_docx(self, path: str) -> str:
        """Extract text from DOCX."""
        doc = docx.Document(path)
        return '\n'.join([para.text for para in doc.paragraphs if para.text])
    
    def _clear_images(self):
        """Clear all selected images."""
        self.image_paths = []
        self.image_preview.clear()
        self.combined_image_preview.clear()
        self._update_mode_ui()
    
    def _clear_document(self):
        """Clear the selected document."""
        self.document_path = None
        self.document_content = None
        self.doc_preview.clear()
        self.combined_doc_preview.clear()
        self.doc_drop.reset()
        self.combined_doc_drop.reset()
        self._update_mode_ui()
    
    def _clear_all(self):
        """Clear all inputs."""
        self._clear_images()
        self._clear_document()
        self.context_input.clear()
        self.results_text.clear()
        self.stats_label.setText("")
        self.export_btn.setEnabled(False)
        self.copy_btn.setEnabled(False)
    
    def _generate_test_cases(self):
        """Start test case generation."""
        # Validate API key
        api_key = self.api_key_input.text().strip()
        if not api_key:
            QMessageBox.warning(self, "API Key Required", "Please enter your Gemini API key.")
            return
        
        # Validate inputs based on mode
        if self.input_mode == "image" and not self.image_paths:
            QMessageBox.warning(self, "Images Required", "Please upload at least one image (screenshot) for Image Only mode.")
            return
        
        if self.input_mode == "document" and not self.document_content:
            QMessageBox.warning(self, "Document Required", "Please upload a document for Document Only mode.")
            return
        
        if self.input_mode == "combined":
            if not self.image_paths and not self.document_content:
                QMessageBox.warning(self, "Input Required", "Please upload at least an image or document for Combined mode.")
                return
        
        # Disable button and show progress
        self.generate_btn.setEnabled(False)
        self.progress_bar.setVisible(True)
        self.progress_bar.setValue(0)
        self.results_text.clear()
        self.results_text.setPlaceholderText("🔄 Generating test cases with Gemini AI...")
        self.status_label.setText("Processing • Generating test cases...")
        
        # Start worker thread
        self.worker = GeminiWorker(
            api_key=api_key,
            image_paths=self.image_paths if self.input_mode in ("image", "combined") else [],
            document_content=self.document_content if self.input_mode in ("document", "combined") else None,
            test_type=self.test_type_combo.currentText(),
            app_context=self.context_input.text(),
            input_mode=self.input_mode
        )
        self.worker.finished.connect(self._on_generation_complete)
        self.worker.error.connect(self._on_generation_error)
        self.worker.progress.connect(self._on_progress)
        self.worker.start()
    
    def _on_progress(self, value: int):
        """Update progress bar."""
        self.progress_bar.setValue(value)
    
    def _on_generation_complete(self, result: str):
        """Handle successful generation."""
        self.results_text.setText(result)
        self.progress_bar.setVisible(False)
        self.generate_btn.setEnabled(True)
        self.export_btn.setEnabled(True)
        self.copy_btn.setEnabled(True)
        
        # Count test cases
        tc_count = result.count("### Test Case")
        timestamp = datetime.now().strftime("%H:%M:%S")
        num_images = len(self.image_paths)
        self.stats_label.setText(f"✓ Generated {tc_count} test cases from {num_images} image(s) • {len(result):,} characters • {timestamp}")
        self.status_label.setText(f"Complete • {tc_count} test cases generated successfully")
    
    def _on_generation_error(self, error: str):
        """Handle generation error."""
        self.progress_bar.setVisible(False)
        self.generate_btn.setEnabled(True)
        QMessageBox.critical(self, "Generation Failed", f"Error: {error}")
        self.results_text.setPlaceholderText(
            "❌ Generation failed. Please check:\n\n"
            "• Your API key is valid\n"
            "• You have API quota remaining\n"
            "• Your input files are readable\n\n"
            "Try again or contact support."
        )
        self.status_label.setText("Error • Generation failed")
    
    def _copy_results(self):
        """Copy results to clipboard."""
        if self.results_text.toPlainText():
            clipboard = QApplication.clipboard()
            clipboard.setText(self.results_text.toPlainText())
            self.status_label.setText("Copied to clipboard!")
            QTimer.singleShot(2000, lambda: self.status_label.setText("Ready"))
    
    def _export_results(self):
        """Export generated test cases."""
        if not self.results_text.toPlainText():
            return
        
        file_path, selected_filter = QFileDialog.getSaveFileName(
            self,
            "Export Test Cases",
            f"android_test_cases_{datetime.now().strftime('%Y%m%d_%H%M%S')}",
            "Excel (*.xlsx);;Markdown (*.md);;Text File (*.txt);;JSON (*.json)"
        )
        
        if not file_path:
            return
        
        content = self.results_text.toPlainText()
        
        try:
            if file_path.endswith('.xlsx'):
                # Excel export
                self._export_to_excel(file_path, content)
            elif file_path.endswith('.json'):
                # Structured JSON export
                data = {
                    "generated_at": datetime.now().isoformat(),
                    "test_type": self.test_type_combo.currentText(),
                    "app_context": self.context_input.text(),
                    "input_mode": self.input_mode,
                    "image_sources": [Path(p).name for p in self.image_paths],
                    "document_source": Path(self.document_path).name if self.document_path else None,
                    "test_cases_raw": content,
                    "test_case_count": content.count("### Test Case")
                }
                with open(file_path, 'w', encoding='utf-8') as f:
                    json.dump(data, f, indent=2, ensure_ascii=False)
            else:
                with open(file_path, 'w', encoding='utf-8') as f:
                    f.write(content)
            
            QMessageBox.information(self, "Export Successful", f"Test cases exported to:\n{file_path}")
            self.status_label.setText(f"Exported to {Path(file_path).name}")
            
        except Exception as e:
            QMessageBox.critical(self, "Export Failed", f"Error: {str(e)}")
    
    def _export_to_excel(self, file_path: str, content: str):
        """Export test cases to Excel with formatting."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Test Cases"
        
        # Define styles
        header_font = Font(bold=True, size=12, color="FFFFFF")
        header_fill = PatternFill(start_color="00897B", end_color="00897B", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        cell_alignment = Alignment(vertical="top", wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Priority colors
        priority_fills = {
            "high": PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid"),
            "medium": PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid"),
            "low": PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
        }
        
        # Headers
        headers = ["TC ID", "Test Case Name", "Priority", "Type", "Preconditions", 
                   "Test Steps", "Expected Results", "Test Data", "Automation Notes"]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Set column widths
        column_widths = [12, 35, 10, 15, 30, 40, 35, 20, 30]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width
        
        # Parse test cases from content
        test_cases = self._parse_test_cases(content)
        
        # Add test case data
        for row_idx, tc in enumerate(test_cases, 2):
            ws.cell(row=row_idx, column=1, value=tc.get("id", "")).border = thin_border
            ws.cell(row=row_idx, column=2, value=tc.get("name", "")).border = thin_border
            
            priority_cell = ws.cell(row=row_idx, column=3, value=tc.get("priority", ""))
            priority_cell.border = thin_border
            priority_lower = tc.get("priority", "").lower()
            if priority_lower in priority_fills:
                priority_cell.fill = priority_fills[priority_lower]
            
            ws.cell(row=row_idx, column=4, value=tc.get("type", "")).border = thin_border
            ws.cell(row=row_idx, column=5, value=tc.get("preconditions", "")).border = thin_border
            ws.cell(row=row_idx, column=6, value=tc.get("steps", "")).border = thin_border
            ws.cell(row=row_idx, column=7, value=tc.get("expected", "")).border = thin_border
            ws.cell(row=row_idx, column=8, value=tc.get("test_data", "")).border = thin_border
            ws.cell(row=row_idx, column=9, value=tc.get("automation", "")).border = thin_border
            
            # Apply alignment to all cells in row
            for col in range(1, 10):
                ws.cell(row=row_idx, column=col).alignment = cell_alignment
        
        # Freeze header row
        ws.freeze_panes = "A2"
        
        # Save workbook
        wb.save(file_path)
    
    def _parse_test_cases(self, content: str) -> List[dict]:
        """Parse test cases from generated content."""
        test_cases = []
        
        # Split by test case markers
        tc_pattern = r'(?:TEST CASE ID:|TC-\d+:|### Test Case)'
        parts = re.split(tc_pattern, content)
        
        for i, part in enumerate(parts[1:], 1):  # Skip first empty part
            tc = {
                "id": f"TC-{i:03d}",
                "name": "",
                "priority": "",
                "type": "",
                "preconditions": "",
                "steps": "",
                "expected": "",
                "test_data": "",
                "automation": ""
            }
            
            lines = part.strip().split('\n')
            current_section = None
            section_content = []
            
            for line in lines:
                line_lower = line.lower().strip()
                line_clean = line.strip()
                
                # Detect sections
                if 'test case name:' in line_lower or line_clean.startswith('TC-'):
                    if ':' in line_clean:
                        tc["name"] = line_clean.split(':', 1)[-1].strip()
                    else:
                        tc["name"] = line_clean
                elif 'priority:' in line_lower or line_lower.startswith('priority'):
                    tc["priority"] = re.sub(r'\*+|priority:?', '', line_clean, flags=re.IGNORECASE).strip()
                elif 'type:' in line_lower and 'test' not in line_lower[:10]:
                    tc["type"] = re.sub(r'\*+|type:?', '', line_clean, flags=re.IGNORECASE).strip()
                elif 'precondition' in line_lower:
                    current_section = "preconditions"
                    section_content = []
                elif 'test step' in line_lower or 'steps:' in line_lower:
                    if current_section and section_content:
                        tc[current_section] = '\n'.join(section_content).strip()
                    current_section = "steps"
                    section_content = []
                elif 'expected result' in line_lower or 'expected:' in line_lower:
                    if current_section and section_content:
                        tc[current_section] = '\n'.join(section_content).strip()
                    current_section = "expected"
                    section_content = []
                elif 'test data' in line_lower:
                    if current_section and section_content:
                        tc[current_section] = '\n'.join(section_content).strip()
                    current_section = "test_data"
                    section_content = []
                elif 'automation' in line_lower:
                    if current_section and section_content:
                        tc[current_section] = '\n'.join(section_content).strip()
                    current_section = "automation"
                    section_content = []
                elif current_section and line_clean and not line_clean.startswith('─') and not line_clean.startswith('═'):
                    # Clean up the line
                    clean_line = re.sub(r'^[\s\-•▸✓\d\.]+', '', line_clean).strip()
                    if clean_line:
                        section_content.append(line_clean)
            
            # Save last section
            if current_section and section_content:
                tc[current_section] = '\n'.join(section_content).strip()
            
            # Only add if we have meaningful content
            if tc["name"] or tc["steps"] or tc["expected"]:
                test_cases.append(tc)
        
        return test_cases


# ═══════════════════════════════════════════════════════════════════════════════
# MAIN ENTRY POINT
# ═══════════════════════════════════════════════════════════════════════════════

def main():
    app = QApplication(sys.argv)
    app.setStyleSheet(STYLESHEET)
    
    # Set application properties
    app.setApplicationName("Android Test Case Generator")
    app.setOrganizationName("TestGen")
    
    window = AndroidTestCaseGenerator()
    window.show()
    
    sys.exit(app.exec())


if __name__ == "__main__":
    main()
